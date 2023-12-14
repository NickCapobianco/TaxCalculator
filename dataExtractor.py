import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Define path and read Excel file
path = "Tax_Calculator.xlsx"
file = pd.read_excel(path, sheet_name="TaxBrackets")

# Data extraction and whitespace/null element handling
data = file.copy()
states = data.loc[0:len(data),'State']
states.dropna(inplace=True)
rates = data.loc[0:len(data),'Rates']
brackets = data.loc[0:len(data),'Brackets']

# Load Excel workbook and worksheet with sheet name specified
wb = load_workbook(filename=path)
ws = wb["Calculator"]

# Reformat the data to save to Excel file
indexCount = 0
ratesList = []
bracketsList = []
tempArr = np.array(states)

for i in range(len(states)):
    tempRatesList = []
    tempBracketsList = []
    tempRatesList.append(tempArr[i])
    tempBracketsList.append(tempArr[i])
    while str(rates.iloc[indexCount]) != 'nan':
        tempRatesList.append(rates.iloc[indexCount])
        tempBracketsList.append(brackets.iloc[indexCount])
        if (indexCount == (len(rates) - 1)):
            break
        indexCount += 1
    tempRatesTuple = (tempRatesList)
    tempBracketsTuple = (tempBracketsList)
    ratesList.append(tempRatesTuple)
    bracketsList.append(tempBracketsTuple)
    indexCount += 1

# Convert data back into dataframes for exporting to Excel
ratesList = pd.DataFrame(ratesList)
bracketsList = pd.DataFrame(bracketsList)
rateCopy = ratesList.copy()
bracketCopy = bracketsList.copy()

# Export data to Excel in a specific column
# Data being exported is state income tax percentages
j = 0
for i in range(len(rateCopy)):
    if rateCopy.iloc[i,0] == ws["B2"].value:
        while str(rateCopy.iloc[i, j + 1]) != 'nan':
            ws[f"C{j + 13}"] = f"=IF(OR($B$2=\"Alaska\", $B$2=\"Florida\", $B$2=\"Nevada\", $B$2=\"New Hampshire\", $B$2=\"South Dakota\", $B$2=\"Tennessee\"," \
                               f" $B$2=\"Texas\", $B$2=\"Washington\", $B$2=\"Wyoming\"),0, {rateCopy.iloc[i, j + 1]})"
            j += 1
            if (j == (len(rateCopy.iloc[i, :]) - 1)):
                break
        while j != (len(rateCopy.iloc[i, :]) - 1):
            ws[f"C{j + 13}"] = f"=IF(OR($B$2=\"Alaska\", $B$2=\"Florida\", $B$2=\"Nevada\", $B$2=\"New Hampshire\", $B$2=\"South Dakota\", $B$2=\"Tennessee\"," \
                                f" $B$2=\"Texas\", $B$2=\"Washington\", $B$2=\"Wyoming\"),0, 0)"
            ws[f"D{j + 13}"] = "=0"
            ws[f"E{j + 13}"] = "N/A"
            j += 1
        break

# Export data to Excel in a specific column
# Data being exported is state income tax brackets
j = 0
for i in range(len(bracketCopy)):
    if bracketCopy.iloc[i,0] == ws["B2"].value:
        while str(bracketCopy.iloc[i, j + 1]) != 'nan':
            if (j >= 1):
                ws[f"D{j + 13}"] = f"=E{j+12}"
                ws[f"E{j + 12}"] = bracketCopy.iloc[i, j + 1]
            else:
                ws[f"D{j + 13}"] = "=0"
            print("Printing value for location [" + str(i) + f",{j}" + "]" + " " + str(ws[f"E{j + 12}"].value))
            j += 1
            if (j == (len(bracketCopy.iloc[i, :]) - 1)):
                break
        while j != (len(bracketCopy.iloc[i, :]) - 1):
            ws[f"E{j + 12}"] = "N/A"
            ws[f"D{j + 13}"] = f"=E{j+12}"
            j += 1
        break


ws = wb["Test"]
wb.save(path)
